#encoding:utf-8

'''
 author     : olenji
 date       : 2016-11-7 14:20:00
 function   : 处理Excel文件的小例子
 py_version ：2.7.2
'''

import re, json, os
#import urllib.parse
#import urllib.request

# #读取excel使用(支持03)
# import xlrd
# #写入excel使用(支持03)
# import xlwt3
#读取execel使用(支持07)
#from openpyxl import workbook
#写入excel使用(支持07)
from openpyxl import load_workbook


class ExcelProcess(object):

    conn = None
    cur = None
    files = []

    #检查是文件还是文件夹，取出.xlsx文件
    def checkInput(self, name):
        if os.path.isfile(name):
            self.files.append(name)
            self.files.append(os.path.dirname(name))
            return True
        if os.path.isfile(name + '.xlsx'):
            self.files.append(os.path.dirname(name + '.xlsx'))
            return True
        if os.path.isdir(name):
            # absname = os.path.dirname(name)
            for dirname, _, filenames in os.walk(name):
                for filename in filenames:
                    if filename.split('.')[-1] == 'xlsx':
                        self.files.append(os.path.join(dirname, filename))
            return True
        else:
            raise NameError("InputError")
            return False


    def getRecords(self, filename):
        #inputWorkbook.get_sheet_names()
        inputWorkbook = load_workbook(filename)
        #for r in xrange(0, work_sheet.get_highest_row()):
        for row in inputWorkbook.worksheets[0]:
            if row[9].value not in [None,'']:
                yield row[1].value, row[9].value


    def record(self, data):
        from openpyxl import Workbook
        from openpyxl.compat import range
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        dest_filename = 'empty_book.xlsx'
        
        ws1 = wb.active
        ws1.title = "range names"
        
        for row in range(1, 40):
            ws1.append([1,2,3,4,5])
        
        ws2 = wb.create_sheet(title="Pi")
        
        ws2['F5'] = 3.14
        
        ws3 = wb.create_sheet(title="Data")
        for row in range(10, 20):
            for col in range(27, 54):
                _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
        print(ws3['AA10'].value)
    
        wb.save(filename = dest_filename)

    def one(self):
        for filename in self.files:
            for a, b in self.getRecords(filename):
                print(a,b)
        pass
if __name__ == '__main__':
    #文件
    #这里需要有文件输入 才能使用
    dirname = 'placedata'
    a = ExcelProcess()
    #a.checkinput(dirname)

#    a.one()
    a.record('a')
    #a.getRecords(os.path.join(dirname,'1.xlsx'))
