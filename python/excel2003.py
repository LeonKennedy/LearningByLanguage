
import re, json, os, pdb
from openpyxl import load_workbook
from openpyxl import Workbook
import xlwt



class AliInputProcess:


    data = dict()
    def __init__(self):
        pass

    def insert(self, q, a):
        if q in self.data.keys():
            self.data[q].append(a)
        else:
            self.data[q] = [a,]

    def getData(self, filename):
        inputWorkbook = load_workbook(filename)
        for row in inputWorkbook.worksheets[1]:
            self.insert(row[1].value.strip(),row[0].value.strip())

    def write2xlsx(self, name):
        wb = Workbook()
        sheet1 = wb.active
        sheet2 = wb.create_sheet(title='sheet2')
        for (q, a)  in self.data.items():
            sheet1.append(a)
            sheet2.append([a[0],q])
        wb.save(name)

    def write2xls(self, name):
        wb = xlwt.Workbook()
        sheet1 = wb.add_sheet('sheet1', cell_overwrite_ok=True)
        sheet2 = wb.add_sheet('sheet2', cell_overwrite_ok=True)
        row = 0
        for (q, a) in self.data.items():
            for i, val in enumerate(a):
                sheet1.write(row, i,val)
            sheet2.write(row, 0, a[0])
            sheet2.write(row, 1, q)
            row += 1
        wb.save(name)

        
    def run(self, filename, destname):
        self.getData(filename)
        self.write2xls(destname)


if __name__ == "__main__":
    a = AliInputProcess()
    filename = 'mike.xlsx'
    a.run(filename, 'a.xls')

